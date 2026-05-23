# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()

# 定义样式
header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # 浅蓝色
header_font = Font(name="微软雅黑", bold=True, size=10)
normal_font = Font(name="微软雅黑", size=10)
red_font = Font(name="微软雅黑", size=10, color="FF0000")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row_num, num_cols):
    """为表头行添加样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row_num, num_cols, red_cols=None, yellow_cols=None):
    """为数据行添加样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
        if red_cols and col in red_cols:
            cell.font = red_font
        if yellow_cols and col in yellow_cols:
            cell.fill = yellow_fill

def auto_adjust_columns(ws, min_width=12, max_width=50):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max(max_length * 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


# ==================== Sheet1: 平台账号速查表 ====================
ws1 = wb.active
ws1.title = "平台账号速查"

# 表头
headers1 = ["平台", "账号名", "账号状态", "ID/绑定信息", "首发日期", "当前待办"]
for col, header in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=header)
style_header_row(ws1, 1, 6)

# 数据
data1 = [
    ["小红书", "潍坊教育观察", "养号中", "小红书号262338479", "5/19", "每天刷15分钟中考教育类内容，点赞关注，不发内容"],
    ["抖音", "（已改名改简介）", "养号中", "新注册号", "5/22", "每天刷15分钟中考教育类视频，点赞关注，不发视频"],
    ["视频号", "潍坊教育观察", "未设置", "用几乎没好友的老微信号", "今天5/16", "今晚：登录→改名字→写简介→发第1条视频"],
    ["知乎", "（待注册）", "未注册", "—", "今天5/16", "今晚：注册→找中考志愿问题→回答1条"],
    ["百度贴吧", "（待确认）", "可直接发", "—", "今天5/16", "今晚：潍坊吧+中考吧各发1条"],
    ["公众号", "（待创建）", "未开", "—", "6月中旬", "等抖音视频号跑起来再开"],
    ["快手", "（待注册）", "未开", "—", "6月下旬", "等前排平台稳定后再开"],
    ["B站", "（待注册）", "未开", "—", "7月初", "等前排平台稳定后再开"],
    ["网站资讯", "—", "运行中", "共10篇文章", "已在更新", "每周更新1-2篇，推GitHub"],
]

for row_num, row_data in enumerate(data1, 2):
    for col, value in enumerate(row_data, 1):
        ws1.cell(row=row_num, column=col, value=value)
    style_data_row(ws1, row_num, 6)

auto_adjust_columns(ws1)


# ==================== Sheet2: 每日工作节奏 ====================
ws2 = wb.create_sheet("每日工作节奏")

headers2 = ["时间段", "名称", "工作内容"]
for col, header in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=header)
style_header_row(ws2, 1, 3)

data2 = [
    ["7:30-8:00", "早间", "刷各平台15分钟（点赞评论养号）+ 回复昨日评论私信"],
    ["12:00-12:30", "午间", "发一条内容（小红书/知乎/贴吧挑一个）+ 回复上午评论"],
    ["20:00-21:00", "晚间", "剪视频/写文案/发视频号抖音 + 次日内容准备"],
]

for row_num, row_data in enumerate(data2, 2):
    for col, value in enumerate(row_data, 1):
        ws2.cell(row=row_num, column=col, value=value)
    style_data_row(ws2, row_num, 3)

auto_adjust_columns(ws2)


# ==================== Sheet3: 分阶段总览 ====================
ws3 = wb.create_sheet("分阶段总览")

headers3 = ["阶段", "时间", "核心目标", "活跃平台"]
for col, header in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=header)
style_header_row(ws3, 1, 4)

data3 = [
    ["第一阶段·养号启动", "5/16-5/22", "养号+铺第一批内容", "小红书/抖音养号，视频号/知乎/贴吧首发"],
    ["第二阶段·内容铺量", "5/23-6/5", "日更节奏建立，每个平台5-10条内容", "全平台铺开（不含快手B站）"],
    ["第三阶段·中考冲刺", "6/6-6/20", "中考热点蹭流量，留资高峰", "全平台高频更新"],
    ["第四阶段·填报指导", "6/21-7/5", "志愿填报干货，转化黄金期", "全平台+公众号启动"],
    ["第五阶段·收网转化", "7/6-7/15", "私域沉淀，报名转化", "全平台引流至公众号/微信"],
]

for row_num, row_data in enumerate(data3, 2):
    for col, value in enumerate(row_data, 1):
        ws3.cell(row=row_num, column=col, value=value)
    style_data_row(ws3, row_num, 4)

auto_adjust_columns(ws3)


# ==================== Sheet4: 每日详细计划 ====================
ws4 = wb.create_sheet("每日详细计划")

headers4 = ["日期", "星期", "时间", "平台", "具体操作", "注意事项"]
for col, header in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=header)
style_header_row(ws4, 1, 6)

data4 = [
    # 5/16（周六）
    ["5/16", "周六", "20:00", "视频号", '用老微信号登录，改名字"潍坊教育观察"，写简介，发1条视频', '简介写"潍坊中考高考志愿填报，关注我不走弯路"，不加微信电话'],
    ["5/16", "周六", "20:30", "知乎", "注册账号，找2-3个中考志愿相关问题，用论坛帖改写回答1条", '回答末尾自然提"我整理了一份志愿填报指南"，引导关注'],
    ["5/16", "周六", "21:00", "百度贴吧", '在"潍坊吧""中考吧"各发1条帖子', "帖子不要带链接，用私信我引导"],
    ["5/16", "周六", "21:15", "小红书", "刷15分钟中考/教育类内容，点赞5-8条，关注3-5个对标号", "不发内容！不评论！不私信！纯刷"],
    ["5/16", "周六", "21:30", "抖音", "刷15分钟中考/教育类视频，点赞5-8条，关注3-5个对标号", "不发视频！不改资料！纯刷"],
    ["5/16", "周六", "21:45", "网站", "确认资讯模块10篇文章正常显示", "如有异常立即修"],
    # 5/17（周日）
    ["5/17", "周日", "7:30", "小红书", "刷10分钟，点赞3-5条", "养号第2天，继续保持"],
    ["5/17", "周日", "12:00", "知乎", "再回答1个中考相关问题", "字数500+，有干货"],
    ["5/17", "周日", "12:30", "百度贴吧", "再发1条不同角度的帖子", "和昨天不要重复内容"],
    ["5/17", "周日", "20:00", "小红书", "刷15分钟，点赞5条，可开始收藏3-5条", "可以收藏了，仍然不发"],
    ["5/17", "周日", "20:20", "抖音", "刷15分钟，点赞5条，收藏3-5条", "不发视频"],
    ["5/17", "周日", "20:40", "视频号", "发第2条视频（脚本第2条改）", "晚8点流量好"],
    ["5/17", "周日", "21:00", "网站", "准备1篇新资讯素材，下周更新用", "素材存备忘"],
    # 5/18（周一）
    ["5/18", "周一", "上午", "域名", "收到阿里云短信后配DNS：CNAME记录wfzgk.com→hh-web-8.github.io，GitHub Pages绑定自定义域名", "没收到短信就等"],
    ["5/18", "周一", "7:30", "小红书", "刷10分钟，点赞评论1-2条", "可以开始评论了，评论要自然"],
    ["5/18", "周一", "12:00", "知乎", "回答1条，或补充之前回答", "持续输出"],
    ["5/18", "周一", "20:00", "小红书", '刷15分钟，评论2-3条（"说得太对了""我家孩子也这样"）', "评论要有真实感，像家长说的"],
    ["5/18", "周一", "20:20", "抖音", "刷15分钟，评论1-2条", "开始互动"],
    ["5/18", "周一", "20:40", "视频号", "发第3条视频", "保持日更"],
    ["5/18", "周一", "21:00", "百度贴吧", "回复自己帖子的评论，顶帖", "贴吧帖子需要维护"],
    # 5/19（周二）⭐小红书首发日
    ["5/19", "周二⭐", "7:30", "小红书", "先刷5分钟，点赞2条", "发之前先活跃一下"],
    ["5/19", "周二⭐", "12:00", "小红书", "⭐发第一篇笔记！（新号版文案）", "不放链接！不留微信！不私信引流！只纯干货"],
    ["5/19", "周二⭐", "12:30", "知乎", "回答1条中考志愿问题", "持续"],
    ["5/19", "周二⭐", "20:00", "视频号", "发第4条视频", ""],
    ["5/19", "周二⭐", "20:20", "小红书", "看第一篇笔记数据，回复评论", "评论必回，增加互动率"],
    ["5/19", "周二⭐", "20:40", "抖音", "继续刷+互动，不发视频", "养号第4天，可改头像和背景图了"],
    ["5/19", "周二⭐", "21:00", "百度贴吧", "发1条新帖或顶帖", ""],
    # 5/20-21（周三周四）
    ["5/20", "周三", "7:30", "小红书", "刷+点赞+评论", "养成习惯"],
    ["5/20", "周三", "12:00", "小红书", "发第2篇笔记", "隔天发一篇，节奏稳定"],
    ["5/20", "周三", "12:30", "知乎", "回答1条", ""],
    ["5/20", "周三", "20:00", "视频号", "发视频（日更）", ""],
    ["5/20", "周三", "20:30", "抖音", "刷+点赞+评论+收藏", "还不发视频"],
    ["5/20", "周三", "21:00", "百度贴吧", "顶帖或新帖", ""],
    ["5/20", "周三", "随时", "网站", "更新1篇资讯到news.json，推送GitHub", "本周至少更新1篇"],
    ["5/21", "周四", "7:30", "小红书", "刷+点赞+评论", "养成习惯"],
    ["5/21", "周四", "12:00", "小红书", "发第3篇笔记", ""],
    ["5/21", "周四", "12:30", "知乎", "回答1条", ""],
    ["5/21", "周四", "20:00", "视频号", "发视频（日更）", ""],
    ["5/21", "周四", "20:30", "抖音", "刷+点赞+评论+收藏", ""],
    ["5/21", "周四", "21:00", "百度贴吧", "顶帖或新帖", ""],
    # 5/22（周五）⭐抖音首发日
    ["5/22", "周五⭐", "7:30", "小红书", "刷+互动", ""],
    ["5/22", "周五⭐", "12:00", "小红书", "发第4篇笔记", ""],
    ["5/22", "周五⭐", "20:00", "抖音", "⭐发第一条视频！（脚本第1条）", "不挂链接！不留微信！一机一卡一号！"],
    ["5/22", "周五⭐", "20:20", "视频号", "发视频", ""],
    ["5/22", "周五⭐", "20:40", "抖音", "看数据，回复评论", ""],
    ["5/22", "周五⭐", "21:00", "知乎", "回答1条", ""],
]

for row_num, row_data in enumerate(data4, 2):
    for col, value in enumerate(row_data, 1):
        ws4.cell(row=row_num, column=col, value=value)
    # 标记关键日期行（⭐标记的）
    if "⭐" in str(row_data[1]):
        style_data_row(ws4, row_num, 6, yellow_cols=[1,2])
    else:
        style_data_row(ws4, row_num, 6)

auto_adjust_columns(ws4, max_width=60)


# ==================== Sheet5: 平台注意事项速查 ====================
ws5 = wb.create_sheet("平台注意事项速查")

headers5 = ["平台", "能用电脑？", "养号天数", "首发日期", "日更频率", "绝对不能做的事"]
for col, header in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=header)
style_header_row(ws5, 1, 6)

data5 = [
    ["小红书", "❌必须手机", "3天（5/17-19）", "5/19", "隔天1篇", "前2周不放链接电话、不私信引流、不互赞互关"],
    ["抖音", "❌必须手机", "7天（5/16-22）", "5/22", "每天1条", "前10条不挂链接、不留微信、一机一卡一号"],
    ["视频号", "❌必须手机", "无需养号", "5/16今天", "每天1条", "新号不要频繁改名字"],
    ["知乎", "✅可以电脑", "无需养号", "5/16今天", "每天回答1条", "不要纯广告，要有干货"],
    ["百度贴吧", "✅可以电脑", "无需养号", "5/16今天", "每周2-3条", "不带链接，用私信我引导"],
    ["公众号", "✅可以电脑", "无需养号", "6月中旬", "每周1-2篇", "不要纯广告，做深度内容"],
    ["快手", "❌必须手机", "无需养号", "6月下旬", "每天1条", "内容要比抖音更接地气"],
    ["B站", "✅可以电脑", "无需养号", "7月初", "每周1条长视频", "要做5-10分钟精品内容"],
    ["网站资讯", "✅可以电脑", "—", "已在更新", "每周1-2篇", "不出现其他技工院校"],
]

for row_num, row_data in enumerate(data5, 2):
    for col, value in enumerate(row_data, 1):
        ws5.cell(row=row_num, column=col, value=value)
    # 绝对不能做的事列（列6）用红色字体
    style_data_row(ws5, row_num, 6, red_cols=[6])

auto_adjust_columns(ws5)


# ==================== Sheet6: 内容素材复用表 ====================
ws6 = wb.create_sheet("内容素材复用表")

headers6 = ["素材源", "小红书", "抖音", "视频号", "知乎", "贴吧", "公众号"]
for col, header in enumerate(headers6, 1):
    ws6.cell(row=1, column=col, value=header)
style_header_row(ws6, 1, 7)

data6 = [
    ["志愿填报指南", "拆成5篇笔记", "提炼成口播脚本", "同抖音", "拆成5个回答", "改成帖子", "全文发"],
    ["抖音视频脚本", "提炼图文要点", "直接用", "直接用", "改成文字回答", "—", "扩展成文章"],
    ["论坛引流帖", "改成图文笔记", "—", "—", "改成回答", "直接用", "—"],
    ["中考政策PDF", "做成信息图笔记", "口播解读", "同抖音", "深度分析", "简版帖子", "完整解读"],
]

for row_num, row_data in enumerate(data6, 2):
    for col, value in enumerate(row_data, 1):
        ws6.cell(row=row_num, column=col, value=value)
    style_data_row(ws6, row_num, 7)

auto_adjust_columns(ws6)


# ==================== Sheet7: 关键时间节点 ====================
ws7 = wb.create_sheet("关键时间节点")

headers7 = ["日期", "事件", "动作"]
for col, header in enumerate(headers7, 1):
    ws7.cell(row=1, column=col, value=header)
style_header_row(ws7, 1, 3)

data7 = [
    ["5/18（周一）", "域名审核预计通过", "配DNS，绑域名，百度站长验证"],
    ["5/19（周二）", "小红书首发", "发第一篇笔记，观察数据"],
    ["5/22（周五）", "抖音首发", "发第一条视频，观察数据"],
    ["6月中旬", "公众号启动", "抖音视频号跑起来后开"],
    ["6月中旬", "中考", "全平台中考热点内容"],
    ["6月下旬", "志愿填报开始", "全平台分数段建议，重点留资"],
    ["6月下旬", "快手启动", "搬运抖音内容"],
    ["7月初", "B站启动", "做精品长视频"],
    ["7/15", "志愿填报截止", "最后冲刺留资"],
]

for row_num, row_data in enumerate(data7, 2):
    for col, value in enumerate(row_data, 1):
        ws7.cell(row=row_num, column=col, value=value)
    # 首发日期用黄色高亮
    if "首发" in str(row_data[1]):
        style_data_row(ws7, row_num, 3, yellow_cols=[1])
    else:
        style_data_row(ws7, row_num, 3)

auto_adjust_columns(ws7)


# 保存文件
output_path = "./招生账号运营日计划.xlsx"
wb.save(output_path)
print(f"Excel文件已生成: {output_path}")
